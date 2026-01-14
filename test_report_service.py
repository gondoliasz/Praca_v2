"""
Test for report_service to verify p-value and recommended test are included in Excel export.
"""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

from services.report_service import generate_excel_report
from openpyxl import load_workbook
from io import BytesIO


def test_excel_report_includes_p_value_and_test():
    """Test that the Excel report includes p-value and recommended test in Summary sheet."""
    
    # Sample result data similar to what r_interface.run_analysis returns
    result = {
        "recommended_test": "pearson_correlation",
        "stats": {
            "method": "Pearson's product-moment correlation",
            "statistic": 2.345,
            "p_value": 0.023456,
            "estimate": 0.567
        },
        "plot_base64": None,
        "actual_x": "Age",
        "actual_y": "Height"
    }
    
    # Generate Excel report
    excel_io = generate_excel_report(result)
    
    # Load the workbook and check the Summary sheet
    wb = load_workbook(excel_io)
    ws = wb["Summary"]
    
    # Convert worksheet to list of rows for easier checking
    rows = list(ws.values)
    
    print("Summary sheet contents:")
    for row in rows:
        print(row)
    
    # Check that recommended test is present
    found_recommended_test = False
    recommended_test_value = None
    for row in rows:
        if row and len(row) >= 2 and row[0] == "Recommended test":
            found_recommended_test = True
            recommended_test_value = row[1]
            break
    
    assert found_recommended_test, "Recommended test field not found in Summary sheet"
    assert recommended_test_value == "pearson_correlation", f"Expected 'pearson_correlation', got '{recommended_test_value}'"
    
    # Check that p-value is present
    found_p_value = False
    p_value_display = None
    for row in rows:
        if row and len(row) >= 2 and row[0] == "P-value":
            found_p_value = True
            p_value_display = row[1]
            break
    
    assert found_p_value, "P-value field not found in Summary sheet"
    assert p_value_display == "0.023456", f"Expected '0.023456', got '{p_value_display}'"
    
    print("\n✓ Test passed: Both recommended test and p-value are present in Summary sheet")


def test_excel_report_handles_zero_p_value():
    """Test that the Excel report correctly handles p-value of 0.0."""
    
    # Sample result with p-value of 0.0
    result = {
        "recommended_test": "t_student",
        "stats": {
            "method": "Student's t-test",
            "statistic": 10.5,
            "p_value": 0.0
        },
        "plot_base64": None,
        "actual_x": "Group",
        "actual_y": "Score"
    }
    
    # Generate Excel report
    excel_io = generate_excel_report(result)
    
    # Load the workbook and check the Summary sheet
    wb = load_workbook(excel_io)
    ws = wb["Summary"]
    
    # Convert worksheet to list of rows
    rows = list(ws.values)
    
    print("\nSummary sheet contents (p-value = 0.0 case):")
    for row in rows:
        print(row)
    
    # Check that p-value of 0.0 is displayed correctly
    found_p_value = False
    p_value_display = None
    for row in rows:
        if row and len(row) >= 2 and row[0] == "P-value":
            found_p_value = True
            p_value_display = row[1]
            break
    
    assert found_p_value, "P-value field not found in Summary sheet"
    assert p_value_display == "0.000000", f"Expected '0.000000' for zero p-value, got '{p_value_display}'"
    
    print("✓ Test passed: P-value of 0.0 is correctly displayed")


if __name__ == "__main__":
    test_excel_report_includes_p_value_and_test()
    test_excel_report_handles_zero_p_value()
    print("\n✅ All tests passed!")
