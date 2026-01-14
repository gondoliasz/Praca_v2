from io import BytesIO
import base64
import json
from typing import Any, Dict
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import pandas as pd

def generate_excel_report(result: Dict[str, Any]) -> BytesIO:
    wb = Workbook()
    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Pole", "Wartość"])
    ws.append(["Recommended test", result.get("recommended_test", "")])
    
    # Extract p-value from stats
    stats = result.get("stats", {})
    p_value_display = ""
    if isinstance(stats, dict):
        p_value_raw = stats.get("p_value")
        if p_value_raw is not None and p_value_raw != "":
            # Format p-value to a reasonable precision
            try:
                p_value_display = f"{float(p_value_raw):.6f}"
            except (ValueError, TypeError):
                p_value_display = str(p_value_raw)
    
    ws.append(["P-value", p_value_display])
    ws.append(["Actual X", result.get("actual_x", "")])
    ws.append(["Actual Y", result.get("actual_y", "")])
    ws.append([])
    ws.append(["Info", "Plik wygenerowany przez aplikację Analiza zależności zmiennych"])

    # Stats sheet - reuse stats from above
    stats_for_sheet = stats if stats else None
    ws2 = wb.create_sheet("Stats")
    if stats_for_sheet is None or (isinstance(stats_for_sheet, (dict, list)) and len(stats_for_sheet) == 0):
        ws2.append(["Brak statystyk"])
    else:
        if isinstance(stats_for_sheet, dict):
            ws2.append(["Nazwa", "Wartość"])
            for k, v in stats_for_sheet.items():
                try:
                    if isinstance(v, (dict, list)):
                        v_str = json.dumps(v, ensure_ascii=False)
                    else:
                        v_str = str(v)
                except Exception:
                    v_str = str(v)
                ws2.append([k, v_str])
        elif isinstance(stats_for_sheet, list):
            try:
                df = pd.DataFrame(stats_for_sheet)
                if df.empty:
                    ws2.append(["Brak danych w liście statystyk"])
                else:
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws2.append(r)
            except Exception:
                ws2.append(["Index", "Value"])
                for i, s in enumerate(stats_for_sheet):
                    ws2.append([i, str(s)])
        else:
            ws2.append([str(stats_for_sheet)])

    # Plot sheet
    plot_b64 = result.get("plot_base64")
    if plot_b64:
        try:
            img_bytes = base64.b64decode(plot_b64)
            pil_img = Image.open(BytesIO(img_bytes)).convert("RGBA")
            img_io = BytesIO()
            pil_img.save(img_io, format="PNG")
            img_io.seek(0)

            ws3 = wb.create_sheet("Plot")
            img_for_xl = OpenpyxlImage(img_io)
            img_for_xl.anchor = "A1"
            ws3.add_image(img_for_xl)
        except Exception:
            ws3 = wb.create_sheet("Plot")
            ws3.append(["Nie udało się załadować wykresu do pliku Excel."])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out